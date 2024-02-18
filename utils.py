# Утилиты
import openpyxl
import inspect
import os


# from core.db_pool import db
from core.db import db
# from core.db_ssh import db

from core import core_log as log


async def save_matrix_to_excel(matrix: list[list]) -> bool:
    """
    Запись матрицы в файл excel
    :param matrix: список списков
    :return:
    """
    workbook = openpyxl.Workbook()
    # список заглавных букв латинского алфавита
    alphabet = [chr(i) for i in range(65, 91)]
    # Оформление границ ячеек.
    # Цвет и стиль границ/бордюров ячеек выставляется атрибутом ячейки .border и классом Border() совместно с классом Side().
    # При этом аргумент стиля границ ячеек border_style может принимать ОДИН из следующих значений: ‘dashDotDot’, ‘medium’, ‘dotted’, ‘slantDashDot’, ‘thin’, ‘hair’, ‘mediumDashDotDot’, ‘dashDot’, ‘double’, ‘mediumDashed’, ‘dashed’, ‘mediumDashDot’ и ‘thick’.
    # определим стили сторон ячеек
    thins = openpyxl.styles.Side(border_style='thin', color='000000')
    double = openpyxl.styles.Side(border_style='double', color='000000')

    # делаем единственный лист активным
    worksheet = workbook.active
    # меняем название листа
    worksheet.title = 'Список товаров'
    worksheet = workbook['Список товаров']
    # Первая строка
    # worksheet.append(['product_id', 'category', 'vendor', 'description', 'price'])
    worksheet.append(['product_id', 'description', 'price'])

    # ширина столбцов листа excel
    worksheet.column_dimensions['A'].width = 13
    # worksheet.column_dimensions['B'].width = 20
    # worksheet.column_dimensions['C'].width = 20
    # worksheet.column_dimensions['D'].width = 60
    # worksheet.column_dimensions['E'].width = 10
    worksheet.column_dimensions['B'].width = 65
    worksheet.column_dimensions['C'].width = 8

    # Установить высоту первой строки
    worksheet.row_dimensions[1].height = 20
    last_col_number = worksheet.max_column

    # номер последней строки
    row_number = worksheet.max_row

    for col_number in range(last_col_number):
        # границы ячейки
        worksheet[f'{alphabet[col_number]}{row_number}'].border = openpyxl.styles.Border(top=thins, bottom=double,
                                                                                         left=thins, right=thins)
        # горизонтальное и вертикальное выравнивания, перенос текста
        worksheet[f'{alphabet[col_number]}{row_number}'].alignment = openpyxl.styles.Alignment(horizontal='left',
                                                                                               vertical='top',
                                                                                               wrap_text=True)
        # шрифт
        worksheet[f'{alphabet[col_number]}{row_number}'].font = openpyxl.styles.Font(name='Arial', size=12, bold=True)
        # заливка
        worksheet[f'{alphabet[col_number]}{row_number}'].fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='F5F5DC')

    # построчное добавление списка в первую свободную строку рабочего листа excel
    for row in matrix:
        worksheet.append(row)
        # номер последней строки
        row_number = worksheet.max_row
        # Установить высоту строки. Метода для автоматической подгонки высоты ячейки нет.
        worksheet.row_dimensions[row_number].height = 13
        for col_number in range(last_col_number):
            # границы ячейки
            worksheet[f'{alphabet[col_number]}{row_number}'].border = openpyxl.styles.Border(top=thins, bottom=thins, left=thins, right=thins)
            # горизонтальное и вертикальное выравнивания, перенос текста
            worksheet[f'{alphabet[col_number]}{row_number}'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=False)
            worksheet[f'C{row_number}'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='top', wrap_text=False)
            # шрифт
            worksheet[f'{alphabet[col_number]}{row_number}'].font = openpyxl.styles.Font(name='Arial', size=10)
            # заливка
            # worksheet[f'D{row_number}'].fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='F0F0F0')
            # форматирование чисел
            worksheet[f'C{row_number}'].number_format = '#,##0'

    # делаем имя файла excel
    filename = 'excel_files/products.xlsx'

    # запись книги в excel-файл
    workbook.save(filename)
    return True


async def save_products() -> None:
    """
    Получение продуктов из БД и вызов функции сохранения матрицы в excel-файл
    :return:
    """
    try:
        # query = 'SELECT product_id, category, vendor, description, price from warehouse WHERE balance > 0 AND price > 0 ORDER BY category, vendor, description'
        query = 'SELECT product_id, description, price from warehouse WHERE balance > 0 AND price > 0 ORDER BY description'
        await db.connect()
        rows = await db.fetch(query=query)
        await db.disconnect()
        # print('Файл готов к записи')
        rows = list(map(list, rows))
        await save_matrix_to_excel(rows)
        # print('Файл записан')
    except Exception as err:
        await log.log(text=f'[no chat_id] {inspect.currentframe().f_code.co_name} {str(err)}', severity='error',
                      facility=os.path.basename(__file__))


if __name__ == '__main__':
    folder = 'excel_files'

    matrix = [['product_id', 'description', 'price'], [17037, 'G. Google PIXEL Tablet with charging speaker dock 128Gb Porcelain wi-fi jp White', 41500]]
    print(save_matrix_to_excel(matrix))

